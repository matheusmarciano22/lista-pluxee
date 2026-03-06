import streamlit as st
import pandas as pd
import re
from datetime import datetime
import io
import os
from unidecode import unidecode
from dateutil import parser
from dateutil.relativedelta import relativedelta
from thefuzz import process
import openpyxl
import requests

# Tenta importar o módulo para leitura de ficheiros Word
try:
    import docx
except ImportError:
    pass

# ==========================================
# LÓGICA DE TRATAMENTO DE DADOS 
# ==========================================

def formatar_nome_pluxee(nome_bruto, limite=40):
    nome = unidecode(str(nome_bruto)).upper().strip()
    nome = re.sub(r'^(NASC|CPF|NOME|VALOR)[:.,\s-]*', '', nome)
    if len(nome) <= limite: return nome
    partes = nome.split()
    if len(partes) <= 2: return nome[:limite]
    primeiro, ultimo = partes[0], partes[-1]
    meio = partes[1:-1]
    for i in range(len(meio)):
        if len(meio[i]) > 2: meio[i] = meio[i][0] + "."
        tentativa = " ".join([primeiro] + meio + [ultimo])
        if len(tentativa) <= limite: return tentativa
    return f"{primeiro} {ultimo}"[:limite]

def formatar_local(texto_bruto, limite=30):
    if pd.isna(texto_bruto): return ""
    texto = unidecode(str(texto_bruto)).upper().strip()
    if len(texto) <= limite: return texto
    partes = texto.split()
    if len(partes) <= 2: return texto[:limite]
    primeiro, ultimo = partes[0], partes[-1]
    meio = partes[1:-1]
    for i in range(len(meio)):
        if len(meio[i]) > 2: meio[i] = meio[i][0] + "."
        tentativa = " ".join([primeiro] + meio + [ultimo])
        if len(tentativa) <= limite: return tentativa
    return texto[:limite]

def limpar_cpf(cpf_bruto):
    if pd.isna(cpf_bruto): return ""
    cpf_limpo = re.sub(r'\D', '', str(cpf_bruto))
    return cpf_limpo.zfill(11)

def converter_data(data_bruta, data_padrao):
    if pd.isna(data_bruta) or str(data_bruta).strip() == "": return data_padrao
    try:
        data_limpa = str(data_bruta).replace('-', '/')
        dt = parser.parse(data_limpa, dayfirst=True, fuzzy=True)
        return dt.strftime('%d/%m/%Y')
    except:
        return data_padrao

def limpar_valor(valor_bruto):
    """Limpa a string de valor para um número decimal puro (ex: 1500.00)."""
    if pd.isna(valor_bruto): return 0
    val_str = str(valor_bruto).upper().replace('R$', '').replace(' ', '').strip()
    if '.' in val_str and ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    elif ',' in val_str:
        val_str = val_str.replace(',', '.')
    try:
        return float(val_str)
    except:
        return 0

# ==========================================
# INTERFACE E LIGAÇÃO API
# ==========================================
st.set_page_config(page_title="Gerador Pluxee Oficial", page_icon="💳", layout="wide")
st.title("💳 Máquina de Geração - Pluxee PLANSIP3C")

col1, col2 = st.columns([1, 2])

if "config_rh" not in st.session_state:
    st.session_state.config_rh = None
if "razao_social" not in st.session_state:
    st.session_state.razao_social = "Cliente_Novo"

with col1:
    st.markdown("### ⚙️ Configurações do Pedido")
    
    tipo_pedido = st.radio("Qual é o Tipo de Pedido?", 
                           ["💳 1ª Via de Cartões (Sem Crédito)", "💰 Recarga de Saldo"], 
                           horizontal=True)
    
    origem_dados = st.radio("De onde deseja puxar os dados do RH (Endereço, etc)?", 
                            ["Lovable (CRM)", "Planilha Antiga do Cliente"], 
                            horizontal=True)
    
    dados_e = {}

    st.markdown("---")

    if origem_dados == "Lovable (CRM)":
        st.markdown("**🔐 Ligação ao Lovable**")
        email_lovable = st.text_input("E-mail Lovable", "seu@email.com")
        senha_lovable = st.text_input("Password Lovable", type="password")
        
        @st.cache_data(ttl=300)
        def carregar_clientes_lovable(email, senha):
            if email == "seu@email.com" or senha == "": return pd.DataFrame()
            url_base = "https://rlbcxlgqfvcloieywfiq.supabase.co"
            anon_key = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJsYmN4bGdxZnZjbG9pZXl3ZmlxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Njk1MzA3NDIsImV4cCI6MjA4NTEwNjc0Mn0.Rj6JhvcGscQ6quQ9K6QwDdiHjbh2pHOZVS7gwDQHDV0"
            try:
                auth = requests.post(f"{url_base}/auth/v1/token?grant_type=password",
                    json={"email": email, "password": senha},
                    headers={"apikey": anon_key, "Content-Type": "application/json"})
                if auth.status_code != 200: return pd.DataFrame()
                token = auth.json()["access_token"]
                resp = requests.get(f"{url_base}/functions/v1/api-vendas",
                    headers={"Authorization": f"Bearer {token}", "apikey": anon_key})
                if resp.status_code == 200:
                    lista = resp.json().get("vendas", [])
                    if lista: return pd.DataFrame(lista).rename(columns={"cliente_nome": "razao_social", "responsavel_pedido": "responsavel", "estado": "uf"})
            except: pass
            return pd.DataFrame()

        df_clientes = carregar_clientes_lovable(email_lovable, senha_lovable)
        
        if not df_clientes.empty:
            st.success(f"✅ {len(df_clientes)} vendas encontradas.")
            empresa_sel = st.selectbox("🔍 Selecione o Cliente", df_clientes['razao_social'].tolist())
            dados_raw = df_clientes[df_clientes['razao_social'] == empresa_sel].iloc[0].to_dict()
            st.session_state.razao_social = empresa_sel
            
            dados_e = {
                'Local de entrega': "MATRIZ",
                'CEP': str(dados_raw.get('endereco_cep', '')),
                'Endereço': str(dados_raw.get('endereco', '')),
                'Número': str(dados_raw.get('numero', '')),
                'Complemento': str(dados_raw.get('endereco_complemento', '')),
                'Bairro': formatar_local(dados_raw.get('endereco_bairro', '')),
                'Cidade': formatar_local(dados_raw.get('cidade', '')),
                'UF': str(dados_raw.get('uf', 'SP')).upper(),
                'Responsável': str(dados_raw.get('responsavel', '')),
                'DDD': "11", 'Telefone': "999999999", 'Email': "", 'Porta_a_Porta': "Não"
            }
        else:
            st.info("Faça login para carregar dados do Lovable.")

    elif origem_dados == "Planilha Antiga do Cliente":
        st.markdown("**📂 Importar Endereço do Histórico**")
        st.session_state.razao_social = st.text_input("Nome da Empresa (Para o arquivo final)", "Cliente_Legado")
        arq_antigo = st.file_uploader("Suba a planilha base da empresa (.xlsx)", type=["xlsx"])
        
        if arq_antigo:
            try:
                wb_antigo = openpyxl.load_workbook(arq_antigo, data_only=True)
                ws_antigo = wb_antigo["Dados dos Beneficiários"]
                l_dados = 8 
                dados_e = {
                    'Local de entrega': str(ws_antigo.cell(row=l_dados, column=16).value or "MATRIZ").replace('None', ''),
                    'CEP': str(ws_antigo.cell(row=l_dados, column=17).value or "").replace('None', ''),
                    'Endereço': str(ws_antigo.cell(row=l_dados, column=18).value or "").replace('None', ''),
                    'Número': str(ws_antigo.cell(row=l_dados, column=19).value or "").replace('None', ''),
                    'Complemento': str(ws_antigo.cell(row=l_dados, column=20).value or "").replace('None', ''),
                    'Bairro': str(ws_antigo.cell(row=l_dados, column=22).value or "").replace('None', '')[:30],
                    'Cidade': str(ws_antigo.cell(row=l_dados, column=23).value or "").replace('None', '')[:30],
                    'UF': str(ws_antigo.cell(row=l_dados, column=24).value or "SP").replace('None', '')[:2].upper(),
                    'Responsável': str(ws_antigo.cell(row=l_dados, column=25).value or "").replace('None', ''),
                    'DDD': str(ws_antigo.cell(row=l_dados, column=26).value or "").replace('None', '')[:2],
                    'Telefone': str(ws_antigo.cell(row=l_dados, column=27).value or "").replace('None', ''),
                    'Email': str(ws_antigo.cell(row=l_dados, column=28).value or "").replace('None', ''),
                    'Porta_a_Porta': str(ws_antigo.cell(row=l_dados, column=29).value or "Não").replace('None', 'Não')
                }
                st.success("✅ Endereço e Responsável capturados com sucesso!")
            except Exception as e:
                st.error("⚠️ Não foi possível ler o arquivo. Tem certeza que é o padrão da Pluxee?")

    st.markdown("**Revisão de Dados do Endereço (Preenchimento Automático)**")
    idx_porta = 1 if "Sim" in str(dados_e.get('Porta_a_Porta', '')) else 0
    st.session_state.config_rh = {
        'Local de entrega': st.text_input("Local de Entrega", dados_e.get('Local de entrega', 'MATRIZ')),
        'CEP': st.text_input("CEP", dados_e.get('CEP', '')),
        'Endereço': st.text_input("Endereço", dados_e.get('Endereço', '')),
        'Número': st.text_input("Número", dados_e.get('Número', '')),
        'Complemento': st.text_input("Complemento", dados_e.get('Complemento', '')),
        'Referência': st.text_input("Referência", ""),
        'Bairro': st.text_input("Bairro", dados_e.get('Bairro', ''), max_chars=30),
        'Cidade': st.text_input("Cidade", dados_e.get('Cidade', ''), max_chars=30),
        'UF': st.text_input("UF", dados_e.get('UF', 'SP'), max_chars=2),
        'Responsável': st.text_input("Responsável", dados_e.get('Responsável', '')),
        'DDD': st.text_input("DDD", dados_e.get('DDD', ''), max_chars=2),
        'Telefone': st.text_input("Telefone", dados_e.get('Telefone', '')),
        'Email': st.text_input("Email", dados_e.get('Email', '')),
        'Porta_a_Porta': st.selectbox("Porta a Porta?", ["Não", "Sim"], index=idx_porta)
    }

with col2:
    st.markdown("### 📥 Lista de Funcionários")
    
    if tipo_pedido == "💰 Recarga de Saldo":
        st.info("💡 **Dica para Recarga:** Envie uma planilha Excel/CSV com as colunas **Nome**, **CPF** e **Valor**.")
    
    # Adicionado suporte ao txt na interface
    arq = st.file_uploader("Suba a lista (Excel, CSV, Word ou TXT)", type=["xlsx", "xls", "csv", "docx", "txt"])
    
    if arq:
        template_path = "PLANSIP3C_NOVA.xlsx"
        if not os.path.exists(template_path):
            st.error("Arquivo PLANSIP3C_NOVA.xlsx não encontrado.")
            st.stop()

        try:
            c_nome = c_cpf = c_nasc = c_valor = None
            
            # --- LEITOR WORD / TXT ---
            if arq.name.endswith('.docx') or arq.name.endswith('.txt'):
                
                # Trata a leitura dependendo se for Word ou TXT
                if arq.name.endswith('.docx'):
                    doc = docx.Document(arq)
                    linhas = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
                else:
                    # Lê o txt ignorando eventuais erros de codificação
                    linhas = [linha.strip() for linha in arq.getvalue().decode("utf-8", errors="ignore").splitlines() if linha.strip()]
                
                dados_ex = []
                p_atual = {'nome': '', 'cpf': '', 'datas': []}
                for l in linhas:
                    nums = re.sub(r'\D', '', l)
                    m_data = re.search(r'\d{2}[/-]\d{2}[/-]\d{2,4}', l)
                    if m_data: p_atual['datas'].append(m_data.group())
                    elif 9 <= len(nums) <= 14:
                        if not p_atual['cpf']: p_atual['cpf'] = nums
                    else:
                        if p_atual['nome'] and p_atual['cpf']:
                            d_nasc = ""
                            if p_atual['datas']:
                                validas = []
                                for d in p_atual['datas']:
                                    try: validas.append((parser.parse(d.replace('-', '/'), dayfirst=True), d))
                                    except: pass
                                if validas:
                                    validas.sort(key=lambda x: x[0])
                                    d_nasc = validas[0][1]
                            dados_ex.append({'nome': p_atual['nome'], 'cpf': p_atual['cpf'], 'nascimento': d_nasc, 'valor': 0})
                            p_atual = {'nome': l, 'cpf': '', 'datas': []}
                        elif not p_atual['nome']: p_atual['nome'] = l
                
                if p_atual['nome'] and p_atual['cpf']:
                    dv = []
                    for d in p_atual['datas']:
                        try: dv.append((parser.parse(d.replace('-', '/'), dayfirst=True), d))
                        except: pass
                    dn = dv[0][1] if dv else ""
                    dados_ex.append({'nome': p_atual['nome'], 'cpf': p_atual['cpf'], 'nascimento': dn, 'valor': 0})
                
                data_rows = pd.DataFrame(dados_ex)
                c_nome, c_cpf, c_nasc, c_valor = 'nome', 'cpf', 'nascimento', 'valor'
                
                if tipo_pedido == "💰 Recarga de Saldo":
                    st.warning("⚠️ Você subiu um Word/TXT para uma Recarga. O sistema não extrai valores soltos de texto. Prefira planilhas (Excel) para Recargas.")

            # --- LEITOR EXCEL/CSV ---
            else:
                if arq.name.endswith('.csv'): df_cli = pd.read_csv(arq, header=None)
                else: df_cli = pd.read_excel(arq, header=None)
                
                start_row = 0
                for i, row in df_cli.head(20).iterrows():
                    l_txt = str(row.values).lower()
                    if 'nome' in l_txt and 'cpf' in l_txt:
                        start_row = i
                        break
                headers = [str(c).lower().strip() for c in df_cli.iloc[start_row]]
                data_rows = df_cli.iloc[start_row+1:].reset_index(drop=True)
                data_rows.columns = headers
                c_nome = headers[headers.index(process.extractOne("nome", headers)[0])]
                c_cpf = headers[headers.index(process.extractOne("cpf", headers)[0])]
                m_nasc = process.extractOne("nascimento", headers)
                c_nasc = headers[headers.index(m_nasc[0])] if m_nasc[1] >= 70 else None
                
                if tipo_pedido == "💰 Recarga de Saldo":
                    m_val = process.extractOne("valor", headers)
                    if m_val and m_val[1] >= 70:
                        c_valor = headers[headers.index(m_val[0])]
                    else:
                        st.warning("⚠️ Coluna 'Valor' não encontrada na planilha.")

            # --- PROCESSAMENTO ---
            if st.session_state.config_rh is not None:
                if st.button("🚀 Gerar Planilha Pluxee Oficial", use_container_width=True):
                    wb = openpyxl.load_workbook(template_path)
                    ws = wb["Dados dos Beneficiários"]
                    r_idx = 8
                    dt_cred = (datetime.now() + relativedelta(months=1)).strftime('%d/%m/%Y')
                    config = st.session_state.config_rh
                    cep_limpo = re.sub(r'\D', '', str(config.get('CEP', '')))
                    
                    if tipo_pedido == "💰 Recarga de Saldo":
                        cod_pedido = "001 - Pedido Normal"
                    else:
                        cod_pedido = "023 - Pedido de 1ªVia de Cartão Sem Crédito"
                    
                    for _, r in data_rows.iterrows():
                        v_n, v_c = r.get(c_nome), r.get(c_cpf)
                        if isinstance(v_n, pd.Series): v_n = v_n.iloc[0]
                        if isinstance(v_c, pd.Series): v_c = v_c.iloc[0]
                        if pd.isna(v_n) or str(v_n).strip() == "" or pd.isna(v_c): continue
                        
                        nf, cpfl = formatar_nome_pluxee(v_n), limpar_cpf(v_c)
                        
                        v_nasc = r.get(c_nasc) if c_nasc else None
                        if isinstance(v_nasc, pd.Series): v_nasc = v_nasc.iloc[0]
                        nasc = converter_data(v_nasc, "01/01/1980") if v_nasc else "01/01/1980"

                        valor_final = 0
                        if tipo_pedido == "💰 Recarga de Saldo" and c_valor:
                            v_val = r.get(c_valor)
                            if isinstance(v_val, pd.Series): v_val = v_val.iloc[0]
                            valor_final = limpar_valor(v_val)

                        for p_code in ["6001 - Carteira Refeição", "6002 - Carteira Alimentação"]:
                            ws.cell(row=r_idx, column=2, value="Ativo")
                            ws.cell(row=r_idx, column=3, value=nf)
                            ws.cell(row=r_idx, column=6, value=nf)
                            ws.cell(row=r_idx, column=4, value=cpfl)
                            ws.cell(row=r_idx, column=5, value=nasc)
                            
                            ws.cell(row=r_idx, column=11, value=cod_pedido)
                            ws.cell(row=r_idx, column=12, value=p_code)
                            ws.cell(row=r_idx, column=13, value=valor_final)
                            
                            ws.cell(row=r_idx, column=14, value=dt_cred)
                            ws.cell(row=r_idx, column=16, value=config.get('Local de entrega'))
                            ws.cell(row=r_idx, column=17, value=cep_limpo)
                            ws.cell(row=r_idx, column=18, value=config.get('Endereço'))
                            ws.cell(row=r_idx, column=19, value=config.get('Número'))
                            ws.cell(row=r_idx, column=20, value=config.get('Complemento'))
                            ws.cell(row=r_idx, column=22, value=config.get('Bairro'))
                            ws.cell(row=r_idx, column=23, value=config.get('Cidade'))
                            ws.cell(row=r_idx, column=24, value=config.get('UF', 'SP'))
                            ws.cell(row=r_idx, column=25, value=config.get('Responsável'))
                            ws.cell(row=r_idx, column=26, value=config.get('DDD'))
                            ws.cell(row=r_idx, column=27, value=config.get('Telefone'))
                            ws.cell(row=r_idx, column=28, value=config.get('Email'))
                            ws.cell(row=r_idx, column=29, value=config.get('Porta_a_Porta'))
                            r_idx += 1
                    
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    st.success("✅ Processado com sucesso!")
                    
                    prefixo = "RECARGA" if tipo_pedido == "💰 Recarga de Saldo" else "PLANSIP3C"
                    st.download_button(label="⬇️ Baixar Planilha Pronto", data=buf, 
                        file_name=f"{prefixo}_{re.sub(r'[^A-Za-z0-9]', '', st.session_state.razao_social)}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.warning("⚠️ Configure os dados da empresa à esquerda primeiro.")
        except Exception as e:
            st.error(f"Erro no processamento: {e}")
